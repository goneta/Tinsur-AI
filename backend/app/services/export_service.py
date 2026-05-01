"""
Export Service – CSV, PDF (ReportLab), and Excel (openpyxl) exports.
"""
import csv
import io
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime

logger = logging.getLogger(__name__)

# ── Optional heavy deps ───────────────────────────────────────────────────────
try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("reportlab not installed – PDF export will be unavailable")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed – Excel export will be unavailable")


class ExportService:

    # ── CSV ───────────────────────────────────────────────────────────────────
    def generate_csv(self, data: List[Dict[str, Any]], headers: List[str] = None) -> str:
        """Generate a CSV string from a list of dictionaries."""
        if not data:
            return ""
        output = io.StringIO()
        if not headers:
            headers = list(data[0].keys())
        writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(data)
        return output.getvalue()

    # ── JSON ──────────────────────────────────────────────────────────────────
    def generate_json(self, data: List[Dict[str, Any]], indent: int = 2) -> str:
        """Serialize data as a pretty-printed JSON string."""
        import json
        return json.dumps(data, default=str, ensure_ascii=False, indent=indent)

    # ── XML ───────────────────────────────────────────────────────────────────
    def generate_xml(self, data: List[Dict[str, Any]], root_name: str = "Records") -> str:
        """Generate XML from a list of dicts. Tag names are sanitized."""
        import re

        def _safe_tag(name: str) -> str:
            tag = re.sub(r"[^a-zA-Z0-9_\-.]", "_", str(name))
            if not tag or tag[0].isdigit():
                tag = "_" + tag
            return tag

        lines = [f'<?xml version="1.0" encoding="UTF-8"?>', f"<{_safe_tag(root_name)}>"]
        for row in data:
            lines.append("  <Record>")
            for k, v in row.items():
                tag = _safe_tag(k)
                val = str(v) if v is not None else ""
                # Escape XML special characters
                val = val.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                lines.append(f"    <{tag}>{val}</{tag}>")
            lines.append("  </Record>")
        lines.append(f"</{_safe_tag(root_name)}>")
        return "\n".join(lines)

    # ── Markdown ──────────────────────────────────────────────────────────────
    def generate_markdown(self, data: List[Dict[str, Any]], title: str = "Export") -> str:
        """Generate a Markdown table from a list of dicts."""
        if not data:
            return f"# {title}\n\n_No data._\n"
        headers = list(data[0].keys())
        lines = [f"# {title}", "", "| " + " | ".join(str(h) for h in headers) + " |",
                 "| " + " | ".join("---" for _ in headers) + " |"]
        for row in data:
            cells = [str(row.get(h, "")).replace("|", "\\|") for h in headers]
            lines.append("| " + " | ".join(cells) + " |")
        lines.append("")
        return "\n".join(lines)

    # ── PDF ───────────────────────────────────────────────────────────────────
    def generate_pdf(self, data: Any, report_type: str) -> bytes:
        """
        Generate a PDF using ReportLab.

        Parameters
        ----------
        data : dict | list
            For tabular reports: a dict with ``title``, ``headers`` (list[str]),
            ``rows`` (list[list]).
            For free-form reports: a dict with ``title`` and ``body`` (str).
        report_type : str
            One of: ``table``, ``policy``, ``quote``, ``claims``, ``financial``.
            Defaults to ``table`` for unknown types.
        """
        if not REPORTLAB_AVAILABLE:
            raise RuntimeError(
                "reportlab is not installed. Run: pip install reportlab"
            )

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2.5 * cm,
            bottomMargin=2 * cm,
        )
        styles = getSampleStyleSheet()
        primary_color = colors.HexColor("#1b5e8c")
        accent_color = colors.HexColor("#0d3b66")

        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Title"],
            fontSize=18,
            textColor=primary_color,
            spaceAfter=6,
        )
        subtitle_style = ParagraphStyle(
            "Subtitle",
            parent=styles["Normal"],
            fontSize=10,
            textColor=colors.HexColor("#666666"),
            spaceAfter=12,
        )
        body_style = ParagraphStyle(
            "Body",
            parent=styles["Normal"],
            fontSize=10,
            leading=14,
        )

        elements: list = []

        # Header
        title = data.get("title", f"Tinsur.AI – {report_type.title()} Report") if isinstance(data, dict) else f"Tinsur.AI Report"
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(
            f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}",
            subtitle_style,
        ))
        elements.append(HRFlowable(width="100%", thickness=1, color=primary_color, spaceAfter=12))

        # Body
        if isinstance(data, dict) and "rows" in data:
            # Tabular report
            headers = data.get("headers", [])
            rows = data.get("rows", [])

            if headers and rows:
                table_data = [headers] + [[str(c) for c in row] for row in rows]
                col_count = len(headers)
                col_width = (A4[0] - 4 * cm) / col_count

                table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
                table.setStyle(TableStyle([
                    # Header row
                    ("BACKGROUND", (0, 0), (-1, 0), primary_color),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("TOPPADDING", (0, 0), (-1, 0), 8),
                    # Data rows
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f5fa")]),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
                    ("TOPPADDING", (0, 1), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ]))
                elements.append(table)

        elif isinstance(data, dict) and "body" in data:
            # Free-form text report
            for line in data["body"].split("\n"):
                elements.append(Paragraph(line or "&nbsp;", body_style))

        elif isinstance(data, list):
            # Auto-table from list of dicts
            if data:
                headers = list(data[0].keys())
                rows = [[str(row.get(h, "")) for h in headers] for row in data]
                return self.generate_pdf(
                    {"title": title, "headers": headers, "rows": rows},
                    report_type,
                )

        # Footer spacer
        elements.append(Spacer(1, 20))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#cccccc")))
        elements.append(Paragraph(
            "Tinsur.AI – Powered by AI  |  Confidential",
            ParagraphStyle("Footer", parent=styles["Normal"], fontSize=7,
                           textColor=colors.HexColor("#aaaaaa"), alignment=TA_CENTER),
        ))

        doc.build(elements)
        return buffer.getvalue()

    # ── Excel ─────────────────────────────────────────────────────────────────
    def generate_excel(
        self,
        data: Any,
        sheet_name: str = "Report",
        title: Optional[str] = None,
    ) -> bytes:
        """
        Generate an Excel (.xlsx) file using openpyxl.

        Parameters
        ----------
        data : list[dict] | dict(with headers+rows)
            Row data for the spreadsheet.
        sheet_name : str
            Name for the worksheet tab.
        title : str | None
            Optional title row inserted at the top.
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError(
                "openpyxl is not installed. Run: pip install openpyxl"
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel tab name limit

        header_fill = PatternFill(start_color="1B5E8C", end_color="1B5E8C", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        alt_fill = PatternFill(start_color="E8F0F7", end_color="E8F0F7", fill_type="solid")
        center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        current_row = 1

        # Optional title row
        if title:
            ws.cell(row=current_row, column=1, value=title).font = Font(
                bold=True, size=13, color="1B5E8C"
            )
            ws.cell(row=current_row, column=1).alignment = center
            ws.cell(row=current_row + 1, column=1, value=
                f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}"
            ).font = Font(italic=True, size=9, color="888888")
            current_row += 3

        # Resolve headers / rows
        if isinstance(data, dict) and "headers" in data:
            headers = data["headers"]
            rows = data.get("rows", [])
        elif isinstance(data, list) and data:
            headers = list(data[0].keys())
            rows = [[row.get(h, "") for h in headers] for row in data]
        else:
            headers = []
            rows = []

        # Header row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=str(header))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Data rows
        for row_idx, row in enumerate(rows):
            fill = alt_fill if row_idx % 2 == 1 else None
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                if fill:
                    cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
            current_row += 1

        # Auto-fit column widths
        for col_idx, header in enumerate(headers, start=1):
            col_values = [str(header)] + [
                str(row[col_idx - 1]) if col_idx - 1 < len(row) else "" for row in rows
            ]
            max_len = min(max((len(v) for v in col_values), default=10), 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
